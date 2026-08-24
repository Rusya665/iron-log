#!/usr/bin/env python3
"""Export Classic Training Split Excel Tracker.

Usage:
    python scripts/export_split_xlsx.py
    python scripts/export_split_xlsx.py --sessions-path path/to/sessions.py
    python scripts/export_split_xlsx.py --out path/to/Latest_Split_Tracker.xlsx
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Ensure iron-log root is on python path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from core.models import Log
from core.plan_generator import calculate_gym_stats
from core.standards import EXERCISE_STANDARDS


def _get_default_paths() -> Tuple[str, str]:
    """Dynamically resolve default paths from local git-ignored config without hardcoding personal folders."""
    sessions_file = os.path.join(PROJECT_ROOT, "sessions.py")
    out_file = os.path.join(PROJECT_ROOT, "dist", "Latest_Split_Tracker.xlsx")

    profiles_path = os.path.join(PROJECT_ROOT, "profiles.json")
    if os.path.exists(profiles_path):
        try:
            with open(profiles_path, "r", encoding="utf-8") as f:
                pdata = json.load(f)
                active_idx = pdata.get("active_profile_index", 0)
                profiles = pdata.get("profiles", [])
                if 0 <= active_idx < len(profiles):
                    p = profiles[active_idx]
                    s_dir = p.get("sessions_dir")
                    o_dir = p.get("output_dir")
                    if s_dir:
                        candidate_s = os.path.join(s_dir, "sessions.py")
                        if os.path.exists(candidate_s):
                            sessions_file = candidate_s
                    if o_dir:
                        out_file = os.path.join(o_dir, "Latest_Split_Tracker.xlsx")
        except Exception:
            pass

    return sessions_file, out_file


def _parse_cli_args() -> argparse.Namespace:
    default_sessions_file, default_out_file = _get_default_paths()

    parser = argparse.ArgumentParser(
        description="Generate classic style Excel tracker for latest split."
    )
    parser.add_argument(
        "--sessions-path",
        type=str,
        default=default_sessions_file,
        help="Path to sessions.py file (default: resolved from local user profile)",
    )
    parser.add_argument(
        "--out",
        type=str,
        default=default_out_file,
        help="Output .xlsx file path (default: resolved from local user profile)",
    )
    parser.add_argument(
        "--weeks",
        type=int,
        default=5,
        help="Minimum number of weeks to scaffold (default: 5)",
    )
    return parser.parse_args()


def load_sessions_module(sessions_path: str):
    """Load sessions.py dynamically as a module."""
    if not os.path.exists(sessions_path):
        raise FileNotFoundError(f"sessions.py not found at {sessions_path}")

    sessions_dir = os.path.dirname(os.path.abspath(sessions_path))
    if sessions_dir not in sys.path:
        sys.path.insert(0, sessions_dir)

    import importlib.util

    spec = importlib.util.spec_from_file_location("sessions", sessions_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Could not load module from {sessions_path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules["sessions"] = mod
    spec.loader.exec_module(mod)
    return mod


def extract_comments(sessions_path: str) -> Dict[str, Dict[str, str]]:
    """Parse comments per exercise and per date from sessions.py."""
    with open(sessions_path, "r", encoding="utf-8", errors="ignore") as f:
        raw_content = f.read()

    # Match each "YYYY-MM-DD": { ... } block
    date_blocks = re.findall(r'"(\d{4}-\d{2}-\d{2})":\s*\{(.*?)\n\s*\}', raw_content, re.DOTALL)
    comments_by_date_ex: Dict[str, Dict[str, str]] = {}

    for d_str, block in date_blocks:
        lines = block.splitlines()
        for line in lines:
            line_clean = line.strip()
            if "#" in line_clean:
                code_part, comment_part = line_clean.split("#", 1)
                comment_text = comment_part.strip().strip('"').strip("'")
                m = re.search(r"([a-zA-Z_]\w*)\s*:\s*Log\(", code_part)
                if m:
                    var_name = m.group(1)
                    comments_by_date_ex.setdefault(d_str, {})[var_name] = comment_text

    return comments_by_date_ex


def format_reps(reps: list) -> Any:
    """Format reps list for display."""
    if not reps:
        return ""
    if len(set(reps)) == 1:
        val = reps[0]
        return int(val) if isinstance(val, (int, float)) and float(val).is_integer() else val
    return ", ".join(str(int(r) if isinstance(r, (int, float)) and float(r).is_integer() else r) for r in reps)


def format_mass(mass: list) -> Any:
    """Format mass list for display."""
    if not mass:
        return ""
    if len(set(mass)) == 1:
        val = mass[0]
        return int(val) if isinstance(val, (int, float)) and float(val).is_integer() else val
    return ", ".join(str(int(m) if isinstance(m, (int, float)) and float(m).is_integer() else m) for m in mass)


def get_display_name(ex_id: str) -> str:
    """Map exercise ID/slug to human-friendly display name."""
    if ex_id in EXERCISE_STANDARDS:
        return EXERCISE_STANDARDS[ex_id].get("name", ex_id.replace("_", " ").replace("-", " ").title())
    cleaned = ex_id.replace("_", " ").replace("-", " ")
    return cleaned.title()


def generate_classic_tracker(sessions_path: str, output_path: str, min_weeks: int = 5) -> str:
    """Generate the classic-style Excel workbook from sessions.py data."""
    # 1. Load user data and extract comments
    mod = load_sessions_module(sessions_path)
    user_data = getattr(mod, "USER_DATA", {})
    comments_by_date_ex = extract_comments(sessions_path)

    # Build var_name -> exercise_id map
    var_to_id: Dict[str, str] = {}
    for k, v in mod.__dict__.items():
        if isinstance(v, str) and not k.startswith("__"):
            var_to_id[k] = v

    # 2. Split Detection
    stats = calculate_gym_stats(user_data)
    split_sessions = stats.get("split_sessions_dates", [])
    cycle_len = stats.get("cycle_length", 3) or 3

    if not split_sessions:
        raise ValueError("No active split sessions found in USER_DATA.")

    # 3. Group split sessions into Weeks (Cycle iterations)
    weeks_sessions: Dict[int, Dict[int, Dict[str, Any]]] = {}
    current_week = 1
    seen_days_in_week = set()

    for s in split_sessions:
        d_str = s["date"]
        d_num = s["day"]

        if d_num in seen_days_in_week:
            current_week += 1
            seen_days_in_week = set()

        seen_days_in_week.add(d_num)
        weeks_sessions.setdefault(current_week, {})[d_num] = {
            "date": d_str,
            "data": user_data.get(d_str, {}),
        }

    total_weeks = max(min_weeks, max(weeks_sessions.keys()) if weeks_sessions else min_weeks)

    # 4. Collect canonical ordered exercises for each Day in the split
    day_ordered_exercises: Dict[int, List[str]] = {}
    for d_num in range(1, cycle_len + 1):
        day_ordered_exercises[d_num] = []

    for s in split_sessions:
        d_num = s["day"]
        d_str = s["date"]
        sess_dict = user_data.get(d_str, {})
        for k, v in sess_dict.items():
            if isinstance(v, Log):
                if k not in day_ordered_exercises[d_num]:
                    day_ordered_exercises[d_num].append(k)

    # 5. Initialize OpenPyXL Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Latest Split Tracker"

    # Styles matching original Rustem_Golden_six.xlsx
    font_header = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)

    border_thin = Side(border_style="thin", color="000000")
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    fill_highlight = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_notes = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # --- TABLE HEADER (Row 1: Weeks with Date Ranges in Brackets) ---
    row_cursor = 1

    for w in range(1, total_weeks + 1):
        col_start = 4 + 4 * (w - 1)
        sess_dict_for_w = weeks_sessions.get(w, {})
        
        # Calculate date range for this week
        if sess_dict_for_w:
            w_dates = [datetime.strptime(info["date"], "%Y-%m-%d") for info in sess_dict_for_w.values() if "date" in info]
            if w_dates:
                d_min = min(w_dates).strftime("%d.%m")
                d_max = max(w_dates).strftime("%d.%m")
                if d_min == d_max:
                    week_label = f"WEEK {w} ({d_min})"
                else:
                    week_label = f"WEEK {w} ({d_min} - {d_max})"
            else:
                week_label = f"WEEK {w}"
        else:
            week_label = f"WEEK {w}"

        ws.cell(row=row_cursor, column=col_start, value=week_label).font = font_header

    ws.row_dimensions[row_cursor].height = 22.0
    row_cursor = 2

    # --- RENDER DAY BLOCKS ---
    for d_num in range(1, cycle_len + 1):
        # 1. Day Header Row
        day_hdr_cell = ws.cell(row=row_cursor, column=2, value=f"Day {d_num}")
        day_hdr_cell.font = font_header
        day_hdr_cell.border = Border(top=border_thin)

        for w in range(1, total_weeks + 1):
            col_start = 4 + 4 * (w - 1)
            headers = ["Sets", "Reps", "Mass", "Notes"]
            for h_idx, h_text in enumerate(headers):
                hc = ws.cell(row=row_cursor, column=col_start + h_idx, value=h_text)
                hc.font = font_regular
                hc.border = Border(top=border_thin)
                hc.alignment = align_center

        ws.row_dimensions[row_cursor].height = 20.0
        row_cursor += 1

        # 2. Exercise Rows
        ex_list = day_ordered_exercises.get(d_num, [])
        for ex_id in ex_list:
            disp_name = get_display_name(ex_id)
            ex_cell = ws.cell(row=row_cursor, column=2, value=disp_name)
            ex_cell.font = font_regular

            for w in range(1, total_weeks + 1):
                col_start = 4 + 4 * (w - 1)
                sess_info = weeks_sessions.get(w, {}).get(d_num)

                sets_c = ws.cell(row=row_cursor, column=col_start + 0)
                reps_c = ws.cell(row=row_cursor, column=col_start + 1)
                mass_c = ws.cell(row=row_cursor, column=col_start + 2)
                notes_c = ws.cell(row=row_cursor, column=col_start + 3)

                for c in (sets_c, reps_c, mass_c, notes_c):
                    c.border = border_cell
                    c.font = font_regular

                sets_c.alignment = align_center
                reps_c.alignment = align_center
                mass_c.alignment = align_center
                notes_c.alignment = align_notes

                if sess_info:
                    d_str = sess_info["date"]
                    day_dict = sess_info["data"]
                    log_obj = day_dict.get(ex_id)

                    if isinstance(log_obj, Log):
                        sets_val = len(log_obj.reps) if log_obj.reps else ""
                        reps_val = format_reps(log_obj.reps)
                        mass_val = format_mass(log_obj.mass)

                        sets_c.value = sets_val
                        reps_c.value = reps_val
                        mass_c.value = mass_val

                        # Find comment
                        note_text = ""
                        date_comments = comments_by_date_ex.get(d_str, {})
                        for vname, ctext in date_comments.items():
                            resolved_id = var_to_id.get(vname, vname)
                            if resolved_id == ex_id or vname.replace("_", "-") == ex_id.replace("_", "-"):
                                note_text = ctext
                                break

                        if note_text:
                            notes_c.value = note_text
                            notes_c.fill = fill_highlight
                            if "+" in note_text or "restored" in note_text or "deload" in note_text:
                                mass_c.fill = fill_highlight

            ws.row_dimensions[row_cursor].height = 20.0
            row_cursor += 1

        row_cursor += 2  # spacing between Day blocks

    # --- COLUMN WIDTHS & PADDING ---
    ws.column_dimensions["A"].width = 3.5
    ws.column_dimensions["B"].width = 28.0
    ws.column_dimensions["C"].width = 3.5

    for w in range(1, total_weeks + 1):
        col_start = 4 + 4 * (w - 1)
        col_sets = get_column_letter(col_start + 0)
        col_reps = get_column_letter(col_start + 1)
        col_mass = get_column_letter(col_start + 2)
        col_notes = get_column_letter(col_start + 3)

        ws.column_dimensions[col_sets].width = 6.0
        ws.column_dimensions[col_reps].width = 7.0
        ws.column_dimensions[col_mass].width = 8.5
        ws.column_dimensions[col_notes].width = 26.0

    out_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
    print(f"[OK] Successfully generated classic split tracker at:\n     {output_path}")
    return output_path


def main():
    args = _parse_cli_args()
    sessions_path = args.sessions_path
    output_path = args.out
    min_weeks = args.weeks

    # Fallback to local sessions.py if path not found
    if not os.path.exists(sessions_path):
        local_candidate = os.path.join(PROJECT_ROOT, "sessions.py")
        if os.path.exists(local_candidate):
            sessions_path = local_candidate

    try:
        out_dir = os.path.dirname(os.path.abspath(output_path))
        os.makedirs(out_dir, exist_ok=True)
    except Exception:
        output_path = os.path.join(PROJECT_ROOT, "dist", "Latest_Split_Tracker.xlsx")

    generate_classic_tracker(sessions_path=sessions_path, output_path=output_path, min_weeks=min_weeks)


if __name__ == "__main__":
    main()
